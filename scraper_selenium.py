import csv

from openpyxl import load_workbook
from selenium import webdriver
from time import sleep

# .xlsx file with company IDs
XLSX_IDS = 'Firmenliste_chid_uid_test.xlsx'

# .csv file to store result
CSV_FILE = 'result.csv'

# Locator for target table data
TARGET_TABLE_LOCATOR = '//table[@class="table table-responsive table-striped personen"]/tbody/tr[contains(@class, "evenRowHideAndSeek") or contains(@class, "oddRowHideAndSeek")]'

# Make generator of IDs from .xslx file
wb = load_workbook(XLSX_IDS)
ws = wb.active
companies_id = ws.iter_rows()

# Request URL
URL = 'https://zh.chregister.ch/cr-portal/auszug/auszug.xhtml?uid='

# Selenium part
driver = webdriver.Chrome('/home/rusrom/webdrivers/chromedriver')

# Iterate through company IDs
for num, row in enumerate(companies_id):

    # Company ID
    uid_format = row[1].value

    # Initialize JSESSION (only once)
    if not num:
        driver.get(URL + uid_format)
        sleep(2)

    # Request page with certain company ID
    driver.get(URL + uid_format)
    sleep(1)

    trs = driver.find_elements_by_xpath(TARGET_TABLE_LOCATOR)

    if num:
        print('Scraping companyID:', uid_format)
    else:
        print('Staring scraping...\n')

    # Scrapig all rows of the target table
    for tr in trs:

        row = [uid_format]
        for i in tr.find_elements_by_tag_name('td'):
            row.append(i.text.strip())

        with open(CSV_FILE, 'a', encoding='utf-8', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(row)

    if not num:
        continue

    print(num, 'company IDs were successfully scraped!\n')

driver.close()
print('SCRAPING COMLETED!')
